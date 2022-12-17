import csv, re, os, datetime
from typing import List, Dict, Tuple, Any
from prettytable import PrettyTable, ALL
import matplotlib
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.workbook import Workbook

import matplotlib.pyplot as plt
import numpy as np
from xlsx2html import xlsx2html

from jinja2 import Environment, FileSystemLoader
import pdfkit

from dateutil.parser import parse


class Report:
    """
    Класс для формирования отчёта, графика и таблицы Excel
    """
    def generate_excel(self, name_vac, statistic):
        """
        Генерирует Excel-файл со статистикой
        Arguments:
             name_vac (str): Название профессии (вакансии)
             statistic (List[Dict[str,str]): Статистика по вакансиям
        """
        thins = Side(border_style="thin", color="000000")
        wb = Workbook()
        sheet1 = wb['Sheet']
        sheet1.title = 'Статистика по годам'
        wb.create_sheet('Статистика по городам')
        columns1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {name_vac}',
                    'Количество вакансий', f'Количество вакансий - {name_vac}']
        for i, column in enumerate(columns1):
            sheet1.cell(row=1, column=(i + 1), value=column).font = Font(bold=True)
        for year, value in statistic[0].items():
            sheet1.append([year, value, statistic[1][year], statistic[2][year], statistic[3][year]])
        for column in sheet1.columns:
            length = max(len(str(cell.value)) for cell in column)
            sheet1.column_dimensions[column[0].column_letter].width = length + 1
            for cell in column:
                cell.border = Border(left=thins, top=thins, right=thins, bottom=thins)
                cell.alignment = Alignment(horizontal='center')
        sheet2 = wb['Статистика по городам']
        columns2 = ['Город', 'Уровень зарплат', '  ', 'Город', 'Доля вакансий']
        for i, column in enumerate(columns2):
            sheet2.cell(row=1, column=(i + 1), value=column).font = Font(bold=True)
        secondStat = list(statistic[4].items())
        for i in range(10):
            secondStat[i] += tuple(statistic[5].items())[i]
        for city1, value1, city2, value2 in secondStat:
            sheet2.append([city1, value1, '  ', city2, value2])
        for i in range(2, 12):
            sheet2[f'E{i}'].number_format = FORMAT_PERCENTAGE_00
        for column in sheet2.columns:
            length = max(len(str(cell.value)) for cell in column)
            sheet2.column_dimensions[column[0].column_letter].width = length + 1
            for cell in column:
                cell.alignment = Alignment(horizontal='center')
                if cell.value != '  ':
                    cell.border = Border(left=thins, top=thins, right=thins, bottom=thins)
        wb.save('report.xlsx')

    def generate_image(self, name_vac, statistic):
        """
            Генерирует png-файл с графиками по статистике
            Arguments:
                name_vac (str): Название профессии (вакансии)
                statistic (List[Dict[str,str]): Статистика по вакансиям
        """
        matplotlib.rc('font', size=8)
        labels = statistic[0].keys()
        total_salaries = statistic[0].values()
        vacancy_salary = statistic[1].values()
        total_count = statistic[2].values()
        vacancy_count = statistic[3].values()
        cities = list(statistic[4].keys())
        cities_salaries = statistic[4].values()
        city_percent = list(statistic[5].values())
        city_percent.insert(0, 1 - sum(city_percent))

        for i in range(len(cities)):
            cities[i] = cities[i].replace(' ', '\n')
            cities[i] = '-\n'.join(cities[i].split('-')) if cities[i].count('-') != 0 else cities[i]

        x = np.arange(len(labels))
        width = 0.35
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)

        ax1.bar(x - width / 2, total_salaries, width, label='средняя з/п')
        ax1.bar(x + width / 2, vacancy_salary, width, label=f'з/п {name_vac}')
        ax1.set_title('Уровень зарплат по годам')
        ax1.set_xticks(x, labels, fontsize=8, rotation=90)
        ax1.legend(loc='upper left', fontsize=8)
        ax1.grid(axis='y')

        ax2.bar(x - width / 2, total_count, width, label='Количество вакансий')
        ax2.bar(x + width / 2, vacancy_count, width, label=f'Количество вакансий {name_vac}')
        ax2.set_title('Количество вакансий по годам')
        ax2.set_xticks(x, labels, fontsize=8, rotation=90)
        ax2.legend(loc='upper left', fontsize=8)
        ax2.grid(axis='y')

        y_pos = np.arange(len(cities))
        ax3.barh(y_pos, cities_salaries, align='center')
        ax3.set_yticks(y_pos, labels=cities, fontsize=6)
        ax3.invert_yaxis()  # labels read top-to-bottom
        ax3.set_title('Уровень зарплат по городам')
        ax3.grid(axis='x')

        x = ['Другие'] + list(statistic[5].keys())
        ax4.set_title('Доля вакансий по городам')
        ax4.pie(city_percent, radius=1, labels=x, textprops={'fontsize': 6})

        fig.tight_layout()
        plt.savefig('graph.png')

    def generate_pdf(self, name_vac):
        """
            Генерирует pdf-файл с Excel и графиками по статистике
            Arguments:
                name_vac (str): Название профессии (вакансии)
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        out1 = xlsx2html('report.xlsx', sheet='Статистика по годам')
        out1.seek(0)
        code1 = out1.read()
        out2 = xlsx2html('report.xlsx', sheet='Статистика по городам')
        out2.seek(0)
        code2 = out2.read()

        pdf_template = template.render({'name_vacancy': name_vac, 'table1': code1, 'table2': code2})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


class DataSet:
    """
        Класс для формирования списка словарей вакансий

        Atributes:
        file_name (str): Имя файла
        vacancies_objects (List[Vacancy]): Список вакансий
        """
    def __init__(self, file_name: str):
        """
        Инициализирует объект Dataset, формирует список вакансий

        arguments:
            file_name (str): Имя файла
        """
        self.file_name = file_name
        self.vacancies_objects = [Vacancy(vac) for vac in self.csv_filer(*self.csv_reader(file_name))]

    def clean_string(self, raw_html) -> str:
        """
        Очищает от html кода
        Arguments:
            raw_html(str): строка для очистки
        returns:
            str: очищенная от html кода строка
        """
        while raw_html.find('<') > -1:
            index1 = raw_html.find('<')
            index2 = raw_html.find('>')
            raw_html = raw_html[:index1] + raw_html[index2 + 1:]
        if '\n' not in raw_html:
            raw_html = " ".join(raw_html.split())
        return raw_html
        # result = re.sub("<.*?>", '', raw_html)
        # return result if '\n' in raw_html else " ".join(result.split())

    def csv_reader(self, file_name: str):
        """
           Считывает данные из файла
            Arguments:
               file_name (str): Название файла
            returns:
                Tuple[List[str], List[List[str]]]: Тюпл, состоящий из списка названий полей и списка списков данных по этим полям
                """

        reader = csv.reader(open(file_name, encoding='utf_8_sig'))
        data_base = [line for line in reader]
        return data_base[0], data_base[1:]

    def csv_filer(self, list_naming, reader) :
        """
            Преобразует данные в список словарей
            Arguments:
                list_naming (List[str]): список полей в вакансиях
                 reader (List[List[str]]): данные с файла

            returns:
                List[Dict[str, str]]: Список словарей, каждый словарь - вакансия
        """
        new_vacans_list = list(filter(lambda vac: (len(vac) == len(list_naming) and vac.count('') == 0), reader))
        return [dict(zip(list_naming, map(self.clean_string, vac))) for vac in new_vacans_list]


class Vacancy:
    """
    Класс для представления вакансий

    Atributes:
    name (str): Название вакансии
    description (str): Описание вакансии
    key_skills (List[str]): Список навыков для вакансии
    experience_id (str): Опыт работы
    premium (str): Премилиальность вакансии
    employer_name (str): Имя работодателя, компании
    salary (Salary): Информация о зарплате
    area_name (str): Город
    published_at (str): Дата публикации вакансии
    """
    def __init__(self, dict_vac):
        """
        Инициализирует объект Vacancy, проверяя наличие полей для Вакансии

        arguments:
            dict_vac (Dict[str,str]): Словарь, хранящий информацию о вакансии (ключ - название поля, значение - характеристика)
        """
        self.name = dict_vac['name']
        self.description = None if 'description' not in dict_vac.keys() else dict_vac['description']
        self.key_skills = None if 'key_skills' not in dict_vac.keys() else dict_vac['key_skills'].split('\n')
        self.experience_id = None if 'experience_id' not in dict_vac.keys() else dict_vac['experience_id']
        self.premium = None if 'premium' not in dict_vac.keys() else dict_vac['premium']
        self.employer_name = None if 'employer_name' not in dict_vac.keys() else dict_vac['employer_name']
        salary_from = dict_vac['salary_from']
        salary_to = dict_vac['salary_to']
        salary_gross = None  if 'salary_gross' not in dict_vac.keys() else dict_vac['salary_gross']
        salary_currency=dict_vac['salary_currency']
        self.salary = Salary(salary_from, salary_to, salary_gross, salary_currency)
        self.area_name = dict_vac['area_name']
        self.published_at = dict_vac['published_at']

class Salary:
    """
    Класс для предсталвения Зарплаты

    Atributes:
    salary_from (str): Нижняя граница зарплаты
    salary_to (str): Высшая граница зарплаты
    salary_gross (str): Наличие/отсутствие налогов
    salary_currency (str): Валюта зарплаты
    """
    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        """
        Инициализирует объект Salary

        arguments:
            salary_from (str): Нижняя граница зарплаты
            salary_to (str): Высшая граница зарплаты
            salary_gross (str): Наличие/отсутствие налогов
            salary_currency (str): Валюта зарплаты
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency

    def to_RUB(self, salary):
        """
        Переводит зарплату в рубли

        arguments:
            salary (float): Зарплата
        returns:
            float: Зарплата в рублях
        >>> Salary('100','1000','false','EUR').to_RUB(559)
        33484.1
        >>> Salary('10', '1000', 'true', 'RUR').to_RUB(500)
        500.0
        >>> Salary('10', '1000', 'true', 'QWE').to_RUB(1000.0)
        Traceback (most recent call last):
        ...
        KeyError: 'QWE'

        """
        return salary * currency_to_rub[self.salary_currency]


class InputConect:
    """
    Класс формирования PrettyTable

    Atributes:
     filter_param (str or List[str]): Параметр фильтрации
     sort_param (str): Параметр сортировки
     reserved_sort (str or bool): Параметр  обратной сортировки
     interval (List[int]): Список двух чисел, задающих интервал строк для вывода таблицы
     columns (str or List[str]): Название выводимых колонок
    """
    def __init__(self, filter_param, sort_param, reversed_sort, interval, columns):
        """
        Инициализирует объект Salary

        arguments:
            filter_param (str): Параметр фильтрации
            sort_param (str): Параметр сортировки
            reserved_sort (str or bool): Параметр  обратной сортировки
            interval (List[int]): Список двух чисел, задающих интервал строк для вывода таблицы
            columns (str or List[str]): Название выводимых колонок
        """
        self.filter_param = filter_param
        self.sort_param = sort_param
        self.reversed_sort = reversed_sort
        self.interval = interval
        self.columns = columns

    def check_parameters(self):
        """
        Проверка параметров на корректность
        """
        if ': ' not in self.filter_param and self.filter_param != '':
            exit_from_file('Формат ввода некорректен')
        self.filter_param = self.filter_param.split(': ')
        if len(self.filter_param) == 2 and self.filter_param[0] not in list(translation.values()):
            exit_from_file('Параметр поиска некорректен')
        if self.sort_param != '' and self.sort_param not in list(translation.values()):
            exit_from_file('Параметр сортировки некорректен')
        if self.reversed_sort not in ['Да', 'Нет', '']:
            exit_from_file('Порядок сортировки задан некорректно')
        self.reversed_sort = (self.reversed_sort == 'Да')
        if len(self.columns) != 0:
            self.columns = self.columns.split(', ')
            self.columns.insert(0, '№')

    def formatter(self,vacancy):
        """
        Осуществляет форматирование необходимых полей

        arguments:
            vacancy (Vacancy): Вакансия

        returns:
            List[any]: список обновленных полей
        """
        def change_salary(salary):
            """
            Форматирует зарплату к нужному формату.

            Arguments:
                salary(Salary): Информация о зарплате
            returns:
                str: Отформатированная информация о зарплате
            """
            salary_from = int(float(salary.salary_from))
            salary_to = int(float(salary.salary_to))
            if salary_from >= 1000:
                salary_from = f'{salary_from // 1000} {str(salary_from)[-3:]}'
                salary_to = f'{salary_to // 1000} {str(salary_to)[-3:]}'
            info_gross = 'Без вычета налогов' if translation[salary.salary_gross] == 'Да' else 'С вычетом налогов'
            result_salary = f'{salary_from} - {salary_to} ({translation[salary.salary_currency]}) ({info_gross})'
            return result_salary

        def change_data(date_vac):
            """
            Форматирует дату к читабельному формату

            Arguments:
                date_vac (str): Дата публикации
            returns:
                str: Отформатированная строка даты
            :return:
            """
            date = date_vac[:date_vac.find('T')].split('-')
            return '.'.join(reversed(date))

        return [vacancy.name, vacancy.description, '\n'.join(vacancy.key_skills), translation[vacancy.experience_id],
                translation[vacancy.premium], vacancy.employer_name, change_salary(vacancy.salary), vacancy.area_name,
                change_data(vacancy.published_at)]

    def data_filter(self,list_vacancies, parameter):
        """
        Фильтрует вакансии по параметру
        Arguments:
             list_vacancies (List[Vacancy]): Список вакансий
             parameter (List[str]): Параметр фильтрации
        returns:
            List[Vacancy]: Отфильтрованный список вакансий
        """
        if parameter[0] == 'Навыки':
            parameter[1] = parameter[1].split(', ')
        if parameter[0] == 'Оклад':
            self.list_vacancies = list(
                filter(lambda vac: int(vac.salary.salary_from) <= int(parameter[1]) <= int(vac.salary.salary_to),
                       list_vacancies))
        elif parameter[0] == 'Навыки':
            list_vacancies = list(
                filter(lambda vac: all(item in vac.key_skills for item in parameter[1]), list_vacancies))
        elif parameter[0] == 'Опыт работы' or parameter[0] == 'Премиум-вакансия':
            list_vacancies = list(
                filter(lambda vac: parameter[1] == translation[vac.__getattribute__(reverse_translation[parameter[0]])],
                       list_vacancies))
        elif parameter[0] == 'Идентификатор валюты оклада':
            list_vacancies = list(
                filter(lambda vac: parameter[1] == translation[vac.salary.salary_currency], list_vacancies))
        elif parameter[0] == 'Дата публикации вакансии':
            list_vacancies = list(
                filter(lambda vac: parameter[1] == '.'.join(reversed(vac.published_at[:vac.published_at.find('T')].split('-')))))
        else:
            list_vacancies = list(
                filter(lambda vac: parameter[1] == vac.__getattribute__(reverse_translation[parameter[0]]),
                       list_vacancies))
        return list_vacancies

    def data_sort(self,list_vacancies, param, is_reverse) :
        """
        Сортирует вакансии по параметру
        Arguments:
            list_vacancies (List[Vacancy]): Список вакансий
            param (List[str]): Параметр фильтрации
            is_reverse(bool): Параметр обратной фильтрации
        returns:
            List[Vacancy]: Отфильтрованный список вакансий
        """
        if param == 'Навыки':
            list_vacancies.sort(key=lambda vac: len(vac.key_skills), reverse=is_reverse)
        elif param == 'Оклад':
            list_vacancies.sort(
                key=lambda vac: vac.salary.to_RUB(float(vac.salary.salary_from) + float(vac.salary.salary_to)) / 2,
                reverse=is_reverse)
        elif param == 'Дата публикации вакансии':
            list_vacancies.sort(
                key=lambda vac: '.'.join(reversed(vac.published_at[:vac.published_at.find('T')].split('-'))),
                reverse=is_reverse)
        elif param == 'Опыт работы':
            list_vacancies.sort(key=lambda vac: rang_experience_id[vac.experience_id], reverse=is_reverse)
        else:
            list_vacancies.sort(key=lambda vac: vac.__getattribute__(reverse_translation[param]), reverse=is_reverse)
        return list_vacancies

    def print_vacancies(self, list_vacancies):
        """
        Печатает вакансии в виде PrettyTable
        Arguments:
            list_vacancies (List[Vacancy]): Список вакансий
        """
        self.interval.append(len(list_vacancies) + 1)
        list_vacancies = list_vacancies if len(self.filter_param) != 2 else self.data_filter(list_vacancies, self.filter_param)
        list_vacancies = list_vacancies if len(list_vacancies) != 0 else 'Ничего не найдено'
        if type(list_vacancies) is str:
            print(list_vacancies)
            return
        list_vacancies = list_vacancies if len(self.sort_param) == 0 else self.data_sort(list_vacancies, self.sort_param, self.reversed_sort)
        table_header = list(reverse_translation.keys())[:-1]
        table_header.insert(0, '№')
        vacans_table = PrettyTable(table_header)
        vacans_table.hrules = ALL
        for i in range(len(list_vacancies)):
            vac = self.formatter(list_vacancies[i])
            vac = list(map(lambda i: f'{i[:100]}...' if len(i) > 100 else i, vac))
            vac.insert(0, i + 1)
            vacans_table.add_row(vac)
        vacans_table.align = 'l'
        vacans_table.max_width = 20
        if len(self.interval) > 1 and len(self.columns) >= 2:
            vacans_table = vacans_table.get_string(start=self.interval[0] - 1, end=self.interval[1] - 1, fields=self.columns)
        elif len(self.interval) > 1:
            vacans_table = vacans_table.get_string(start=self.interval[0] - 1, end=self.interval[1] - 1)
        elif len(self.columns) >= 2:
            vacans_table = vacans_table.get_string(fields=self.columns)
        print(vacans_table)


def get_salary_level(list_vacancies, field, name_vacancy=''):
    """
    Формирует статистики, связанные с зарплатами
    Atributes:
        list_vacancies (List[Vacancy]): Список вакансий
        field (str): Поле вакансии
        name_vacancy (str): Название вакансии(если ее ввели)
    returns:
        Dict[str,str]:  Статистики, связанные с зарплатой
    """
    result = {}
    for vac in list_vacancies:
        result[vac.__getattribute__(field)] = [] if vac.__getattribute__(field) not in result.keys() else result[vac.__getattribute__(field)]
    list_vacancies = list_vacancies if name_vacancy == '' else list(filter(lambda vac: name_vacancy in vac.name, list_vacancies))
    for vac in list_vacancies:
        result[vac.__getattribute__(field)].append(vac.salary.to_RUB(float(vac.salary.salary_from) + float(vac.salary.salary_to)) / 2)
    for key in result.keys():
        result[key] = 0 if len(result[key]) == 0 else int(sum(result[key]) // len(result[key]))
    return result


def get_count_vacancies(list_vacancies, field, name_vacancy=''):
    """
    Формирует статистики, связанные с количеством вакансий
    Atributes:
        list_vacancies (List[Vacancy]): Список вакансий
        field (str): Поле вакансии
        name_vacancy (str): Название вакансии(если ее ввели)
    returns:
        Dict[str,str]:  Статистики, связанные с количеством вакансий
    """
    result = {}
    for vac in list_vacancies:
        result[vac.__getattribute__(field)] = 0 if vac.__getattribute__(field) not in result.keys() else result[vac.__getattribute__(field)]
    list_vacancies = list_vacancies if name_vacancy == '' else list(filter(lambda vac: name_vacancy in vac.name, list_vacancies))
    for vac in list_vacancies:
        result[vac.__getattribute__(field)] += 1
    if field == 'area_name':
        for key in result.keys():
            result[key] = round(result[key] / len(data.vacancies_objects), 4)
    return result


currency_to_rub = {"AZN": 35.68,
                   "BYR": 23.91,
                   "EUR": 59.90,
                   "GEL": 21.74,
                   "KGS": 0.76,
                   "KZT": 0.13,
                   "RUR": 1,
                   "UAH": 1.64,
                   "USD": 60.66,
                   "UZS": 0.0055, }

translation = {"name": "Название",
               "description": "Описание",
               "key_skills": "Навыки",
               "experience_id": "Опыт работы",
               "premium": "Премиум-вакансия",
               "employer_name": "Компания",
               "salary_from": "Нижняя граница вилки оклада",
               "salary_to": "Верхняя граница вилки оклада",
               "salary_gross": "Оклад указан до вычета налогов",
               "salary_currency": "Идентификатор валюты оклада",
               "area_name": "Название региона",
               "published_at": "Дата публикации вакансии",
               "Оклад": "Оклад",
               "True": "Да",
               "TRUE": "Да",
               "False": "Нет",
               "FALSE": "Нет",
               "noExperience": "Нет опыта",
               "between1And3": "От 1 года до 3 лет",
               "between3And6": "От 3 до 6 лет",
               "moreThan6": "Более 6 лет",
               "AZN": "Манаты",
               "BYR": "Белорусские рубли",
               "EUR": "Евро",
               "GEL": "Грузинский лари",
               "KGS": "Киргизский сом",
               "KZT": "Тенге",
               "RUR": "Рубли",
               "UAH": "Гривны",
               "USD": "Доллары",
               "UZS": "Узбекский сум"}

reverse_translation = {"Название": "name",
                       "Описание": "description",
                       "Навыки": "key_skills",
                       "Опыт работы": "experience_id",
                       "Премиум-вакансия": "premium",
                       "Компания": "employer_name",
                       "Оклад": "Оклад",
                       "Название региона": "area_name",
                       "Дата публикации вакансии": "published_at",
                       "Идентификатор валюты оклада": "salary_currency"}

rang_experience_id = {"noExperience": 0,
                      "between1And3": 1,
                      "between3And6": 2,
                      "moreThan6": 3}

def change_data(date_vac) -> str:
    """
    Форматирует дату публикации к нужному формату

    Args:
        date_vac(str): Дата публикации
    returns:
        str: Отформатированная дата публикации
    """
    return date_vac[:4]

def exit_from_file(message):
    """
    Прерывает программу выполнения, выводя перд этим сообщение
    Arguments: message(str): Сообщение
    """
    print(message)
    exit()

def get_statistic(result_list, index, is_reversed=False, slice=0):
    """

    Преобразовывает статистику
    Arguments:
        result_list (Dict[Any,Any]): Статистика
        index (int): Индекс
        is_reversed(bool): Фильтр сортировки
        slice (int): Срез статистики
    returns:
        Dict[Any,Any]: Словарь преобразованной статистики
    """
    slice = len(result_list) if slice == 0 else slice
    return dict(sorted(result_list, key=lambda x: x[index], reverse=is_reversed)[:slice])

type_output=input('Введите данные для печати: ')
file_name = input('Введите название файла: ')
if os.stat(file_name).st_size == 0:
    exit_from_file('Пустой файл')
data = DataSet(file_name)
if len(data.vacancies_objects) == 0:
    exit_from_file('Нет данных')
if type_output=='Статистика':
    vacancy_name = input('Введите название профессии: ')
    for vac in data.vacancies_objects:
        vac.published_at = change_data(vac.published_at)
    dict_cities = {}
    for vac in data.vacancies_objects:
        if vac.area_name not in dict_cities.keys():
            dict_cities[vac.area_name] = 0
        dict_cities[vac.area_name] += 1
    needed_vacancies_objects = list(filter(lambda vac: int(len(data.vacancies_objects) * 0.01) <= dict_cities[vac.area_name], data.vacancies_objects))

    rp = Report()
    list_statistic = [get_statistic(get_salary_level(data.vacancies_objects, 'published_at').items(), 0),
                  get_statistic(get_salary_level(data.vacancies_objects, 'published_at', vacancy_name).items(), 0),
                  get_statistic(get_count_vacancies(data.vacancies_objects, "published_at").items(), 0),
                  get_statistic(get_count_vacancies(data.vacancies_objects, 'published_at', vacancy_name).items(), 0),
                  get_statistic(get_salary_level(needed_vacancies_objects, 'area_name').items(), 1, True, 10),
                  get_statistic(get_count_vacancies(needed_vacancies_objects, 'area_name').items(), 1, True, 10)]
    rp.generate_excel(vacancy_name, list_statistic)
    rp.generate_image(vacancy_name, list_statistic)
    rp.generate_pdf(vacancy_name)
elif 'Вакансии':
    parametrOfFiltr = input('Введите параметр фильтрации: ')
    parametrOfSort = input('Введите параметр сортировки: ')
    isReversedSort = input('Обратный порядок сортировки (Да / Нет): ')
    rowsNumbers = list(map(int,input('Введите диапазон вывода: ').split()))
    collomnNames = input('Введите требуемые столбцы: ')
    outer=InputConect(parametrOfFiltr, parametrOfSort,isReversedSort,rowsNumbers,collomnNames)
    outer.check_parameters()
    outer.print_vacancies(data.vacancies_objects)