import csv, os, re,datetime, openpyxl

import matplotlib
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

import matplotlib.pyplot as plt
import numpy as np

class Report:
    def createExcel(self, name_vac, statistic):
        thins = Side(border_style="thin", color="000000")
        wb = Workbook()
        sheet1 = wb['Sheet']
        sheet1.title = 'Статистика по годам'
        wb.create_sheet('Статистика по городам')
        columns1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {name_vac}',
                    'Количество вакансий', f'Количество вакансий - {name_vac}']
        for i, column in enumerate(columns1):
            sheet1.cell(row=1, column=(i + 1), value=column).font = Font(bold=True)
        for year, value in statistic[0].items():
            sheet1.append([year, value, statistic[1][year], statistic[2][year], statistic[3][year]])
        for column in sheet1.columns:
            length = max(len(str(cell.value)) for cell in column)
            sheet1.column_dimensions[column[0].column_letter].width = length + 1
            for cell in column:
                cell.border = Border(left=thins, top=thins, right=thins, bottom=thins)
        sheet2 = wb['Статистика по городам']
        columns2 = ['Город', 'Уровень зарплат', '  ', 'Город', 'Доля вакансий']
        for i, column in enumerate(columns2):
            sheet2.cell(row=1, column=(i + 1), value=column).font = Font(bold=True)
        # secondStat = list(tuple(statistic[4].items()) + tuple(statistic[5].items()))
        secondStat = list(statistic[4].items())
        for i in range(10):
            secondStat[i] += tuple(statistic[5].items())[i]
        for city1, value1, city2, value2 in secondStat:
            sheet2.append([city1, value1, '  ', city2, value2])
        for i in range(2, 12):
            sheet2[f'E{i}'].number_format = FORMAT_PERCENTAGE_00
        for column in sheet2.columns:
            length = max(len(str(cell.value)) for cell in column)
            sheet2.column_dimensions[column[0].column_letter].width = length + 1
            for cell in column:
                if cell.value != '  ':
                    cell.border = Border(left=thins, top=thins, right=thins, bottom=thins)
        wb.save('report.xlsx')

    def generate_image(self, name_vac, statistic):
        matplotlib.rc('font', size=8)
        labels = statistic[0].keys()
        salaryYears = statistic[0].values()
        salaryYearsJob = statistic[1].values()
        vacancyYears=statistic[2].values()
        vacancyYearsJob=statistic[3].values()
        citySalaries=statistic[4].values()
        cityVacancies=list(statistic[5].values())
        cityVacancies=[1-sum(cityVacancies)]+cityVacancies
        cities=list(statistic[4].keys())

        for i in range (len(cities)):
            cities[i]=cities[i].replace(' ','\n')
            cities[i] = '-\n'.join(cities[i].split('-')) if cities[i].count('-') != 0 else cities[i]

        x = np.arange(len(labels))  # the label locations
        width = 0.35  # the width of the bars

        fig, ((ax1,ax2),(ax3,ax4)) = plt.subplots(nrows=2, ncols=2)

        #1
        ax1.bar(x - width / 2, salaryYears, width, label='средняя з/п')
        ax1.bar(x + width / 2, salaryYearsJob, width, label='з/п '+name_vac.lower())
        ax1.set_title('Уровень зарплат по годам')
        ax1.set_xticks(x, labels, fontsize=8, rotation=90)
        ax1.legend(loc='upper left', fontsize=8)
        ax1.grid(axis='y')

        #2
        ax2.bar(x - width / 2, vacancyYears, width, label='Количество вакансий')
        ax2.bar(x + width / 2, vacancyYearsJob, width, label='Количество вакансий ' + name_vac.lower())
        ax2.set_title('Количество вакансий по годам')
        ax2.set_xticks(x, labels, fontsize=8, rotation=90)
        ax2.legend(loc='upper left', fontsize=8)
        ax2.grid(axis='y')

        #3
        y_pos = np.arange(len(cities))
        performance = citySalaries
        ax3.barh(y_pos, performance, align='center')
        ax3.set_yticks(y_pos, labels=cities, fontsize=6)
        ax3.invert_yaxis()
        ax3.set_title('Уровень зарплат по городам')
        ax3.grid(axis='x')

        #4
        citiesOther=['Другие']+list(statistic[5].keys())
        ax4.set_title('Доля вакансий по городам')
        ax4.pie(cityVacancies, radius=1,labels=citiesOther, textprops={'fontsize':6})
        fig.tight_layout()

        plt.show()

class DataSet:
    def __init__(self, file_name):
        self.file_name=file_name
        self.vacancies_objects=[Vacancy(vac) for vac in self.csv_filer(*self.csv_reader(file_name))]

    def csv_reader(self,file):
        listFile = []
        with open(file, encoding='utf-8-sig') as File:
            reader = csv.reader(File, delimiter=',')
            for row in reader:
                listFile.append(row)
        return listFile[1:], listFile[0]

    def csv_filer(self,reader, list_naming):
        listFile = []
        for item in reader:
            if item.count('') == 0 and len(item) == len(list_naming):
                listFile.append(item)
        allVacancies = []
        for i in range(len(listFile)):
            vacancy = {}
            for j in range(len(list_naming)):
                name = list_naming[j]
                data = listFile[i][j]
                while data.find('<') > -1:
                    index1 = data.find('<')
                    index2 = data.find('>')
                    data = data[:index1] + data[index2 + 1:]
                if '\n' not in data:
                    data = " ".join(data.split())
                vacancy[name] = data
            allVacancies.append(vacancy)
        return allVacancies

class Vacancy:
    def __init__(self,vacancy_dict):
        self.name=vacancy_dict['name']
        self.salary=Salary(vacancy_dict['salary_from'], vacancy_dict['salary_to'], vacancy_dict['salary_currency'])
        self.area_name=vacancy_dict['area_name']
        self.published_at=vacancy_dict['published_at']

class Salary:
    def __init__(self,salary_from,salary_to,salary_currency):
        self.salary_from=salary_from
        self.salary_to=salary_to
        self.salary_currency=salary_currency

    def toRub(self, salary):
        return salary * currency_to_rub[self.salary_currency]

class InputConect:
    def get_salary_level(self,vacancyList, key, vacancyName=''):
        vacancyDict={}
        for vacancy in vacancyList:
            newKey = vacancy.__getattribute__(key)
            if newKey not in vacancyDict.keys():
                vacancyDict[newKey] = []
        if vacancyName!='':
            vacancyList = list(filter(lambda vacancy: vacancyName in vacancy.name, vacancyList))
        for vacancy in vacancyList:
            newKey=vacancy.__getattribute__(key)
            vacancyDict[newKey].append(vacancy.salary.toRub(float(vacancy.salary.salary_from)+float(vacancy.salary.salary_to))/2)
        for keyy in vacancyDict.keys():
            if len(vacancyDict[keyy])==0:
                vacancyDict[keyy]=0
                continue
            vacancyDict[keyy]=int(sum(vacancyDict[keyy]) // len(vacancyDict[keyy]))
        return vacancyDict

    def get_vacancy_amount(self, vacancyList, key,dataset,vacancyName=''):
        vacancyDict = {}
        for vacancy in vacancyList:
            newKey = vacancy.__getattribute__(key)
            if newKey not in vacancyDict.keys():
                vacancyDict[newKey] = 0
        if vacancyName != '':
            vacancyList = list(filter(lambda vacancy: vacancyName in vacancy.name, vacancyList))
        for vacancy in vacancyList:
            newKey = vacancy.__getattribute__(key)
            vacancyDict[newKey]+=1
        if key=='area_name':
            for keyy in vacancyDict.keys():
                vacancyDict[keyy]=round(vacancyDict[keyy] /len(dataset.vacancies_objects),4)
        return vacancyDict

def get_year_of_date(date_vac):
    return datetime.datetime.strptime(date_vac, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y')

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

def input_data():
    file = input('Введите название файла: ')
    job = input('Введите название профессии: ')
    if os.stat(file).st_size==0:
        print("Пустой файл")
        exit()
    dataset = DataSet(file)
    if len(dataset.vacancies_objects)==0:
        print('Нет данных')
        exit()
    outer = InputConect()
    for vacancy in dataset.vacancies_objects:
        vacancy.published_at=int(get_year_of_date(vacancy.published_at))
    som1=dict(sorted(outer.get_salary_level(dataset.vacancies_objects,"published_at").items(),key=lambda x:int(x[0])))
    print(f'Динамика уровня зарплат по годам: {som1}')
    som2=dict(sorted(outer.get_vacancy_amount(dataset.vacancies_objects,"published_at",dataset).items(),key=lambda x:int(x[0])))
    print(f'Динамика количества вакансий по годам: {som2}')
    som3 = dict(sorted(outer.get_salary_level(dataset.vacancies_objects, "published_at",job).items(), key=lambda x: int(x[0])))
    print(f'Динамика уровня зарплат по годам для выбранной профессии: {som3}')
    som4 = dict(sorted(outer.get_vacancy_amount(dataset.vacancies_objects, "published_at",dataset,job).items(), key=lambda x: int(x[0])))
    print(f'Динамика количества вакансий по годам для выбранной профессии: {som4}')

    dict_cities={}
    for vac in dataset.vacancies_objects:
        if vac.area_name not in dict_cities.keys():
            dict_cities[vac.area_name]=0
        dict_cities[vac.area_name]+=1
    otherVacanciesObjects=list(filter(lambda vacancy: int(len(dataset.vacancies_objects)*0.01)<= dict_cities[vacancy.area_name], dataset.vacancies_objects))

    som5 = dict(sorted(outer.get_salary_level(otherVacanciesObjects, "area_name").items(), key=lambda x: x[1], reverse=True)[:10])
    print(f'Уровень зарплат по городам (в порядке убывания): {som5}')
    som6 = dict(sorted(outer.get_vacancy_amount(otherVacanciesObjects, "area_name",dataset).items(), key=lambda x: x[1], reverse=True)[:10])
    print(f'Доля вакансий по городам (в порядке убывания): {som6}')

    rp=Report()
    print(som1)
    statistic=[som1,som3,som2,som4,som5,som6]
    print(statistic)
    rp.createExcel(job, statistic)
    rp.generate_image(job, statistic)
input_data()
